from google.oauth2 import service_account
from googleapiclient.discovery import build
import openpyxl
from io import BytesIO


def get_filenames():
    """
    Получает список файлов из указанной папки на Google Диске.

    Возвращает:
        list: Список словарей файлов, содержащих идентификаторы и имена файлов.
    """
    scopes = ["https://www.googleapis.com/auth/drive"]
    SERVICE_ACCOUNT_FILE = (
        "/home/cusdeb/Projects/lizard_bot/lizardbot/lizardbot-423509-18b41a862983.json"
    )
    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=scopes
    )

    drive_service = build("drive", "v3", credentials=credentials)

    folder_id = "19yyXXullGGMIT3XISiZ33wkDxHJy0zvb"
    results = (
        drive_service.files()
        .list(q=f"'{folder_id}' in parents", fields="nextPageToken, files(id, name)")
        .execute()
    )
    files = results.get("files", [])

    return files


def download_file(file_id, drive_service):
    """
    Скачивает файл с Google Диска.

    Аргументы:
        file_id (str): Идентификатор файла для скачивания.
        drive_service (googleapiclient.discovery.Resource): Экземпляр службы Google Диска.

    Возвращает:
        BytesIO: Содержимое скачанного файла.
    """
    request = drive_service.files().get_media(fileId=file_id)
    file_content = BytesIO(request.execute())
    return file_content


def process_excel(file_content, group_name):
    """
    Обрабатывает Excel файл для извлечения информации о расписании для указанной группы.

    Аргументы:
        file_content (BytesIO): Содержимое Excel файла.
        group_name (str): Название группы для поиска.

    Возвращает:
        list: Список кортежей с информацией о расписании (название листа, номер кабинета, имя преподавателя, флаг классного часа).
    """
    wb = openpyxl.load_workbook(file_content)
    results = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for i in range(1, len(row), 3):
                if row[i] == group_name:
                    room_number = row[i - 1]
                    teacher_name = row[i + 1]
                    class_hour = "Классный час" in row
                    results.append((sheet.title, room_number, teacher_name, class_hour))
    return results


def process_excel2(file_content, teacher_name):
    """
    Обрабатывает Excel файл для извлечения информации о расписании для указанного преподавателя.

    Аргументы:
        file_content (BytesIO): Содержимое Excel файла.
        teacher_name (str): Имя преподавателя для поиска.

    Возвращает:
        list: Список кортежей с информацией о расписании (название листа, номер кабинета, название группы, полное имя преподавателя, флаг классного часа).
    """
    wb = openpyxl.load_workbook(file_content)
    results = []
    teacher_last_name = teacher_name.strip().lower()

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for i in range(2, len(row), 3):
                if row[i] and row[i].strip().lower().startswith(teacher_last_name):
                    room_number = row[i - 2]
                    group_name = row[i - 1]
                    full_teacher_name = row[i]
                    class_hour = "Классный час" in row
                    results.append(
                        (
                            sheet.title,
                            room_number,
                            group_name,
                            full_teacher_name,
                            class_hour,
                        )
                    )
    return results


def form_schedule(schedule):
    """
    Форматирует расписание, добавляя время для каждой пары.

    Аргументы:
        schedule (str): Строка с расписанием.

    Возвращает:
        str: Отформатированное расписание.
    """
    time_mapping = {
        1: "8:00 - 9:30",
        2: "9:40 - 11:10",
        3: "11:30 - 13:00",
        4: "13:10 - 14:40",
        5: "15:00 - 16:30",
        6: "16:40 - 18:10",
        7: "18:20 - 19:50",
    }
    extended_time_mapping = {
        1: "8:00 - 9:30",
        2: "9:40 - 11:10",
        3: "11:30 - 13:00",
        4: "14:10 - 15:40",
        5: "16:00 - 17:30",
        6: "17:40 - 19:10",
    }
    schedule_text = schedule.split("\n")
    class_hour_day = any("Классный час" in line for line in schedule_text)

    for i, line in enumerate(schedule_text):
        schedule_text_str = ""
        if line.strip() and line[0].isdigit():
            pair_number = int(line[0])
            if class_hour_day and pair_number in time_mapping:
                schedule_text_str = (
                    "🕒 " + schedule_text[i] + f" {extended_time_mapping[pair_number]}"
                )
            elif pair_number in time_mapping:
                schedule_text_str = (
                    "🕒 " + schedule_text[i] + f" {time_mapping[pair_number]}"
                )
            schedule_text[i] = schedule_text_str
        elif "Классный час" in line:
            schedule_text[i] = "🕒 " + line
    updated_schedule = "\n".join(schedule_text)
    return updated_schedule


def service(name, group):
    """
    Обрабатывает запрос на получение расписания для указанной группы.

    Аргументы:
        name (str): Название файла с расписанием (без расширения).
        group (str): Название группы для поиска.

    Возвращает:
        str: Отформатированное расписание или сообщение об ошибке.
    """
    files = get_filenames()

    for index, file in enumerate(files):
        print(f"{index + 1}. {file['name']}")

    chosen_file_name = name

    chosen_file = None
    for file in files:
        if file["name"] == chosen_file_name + ".xlsx":
            chosen_file = file
            break

    if chosen_file is None:
        print("Файл не найден.")
        return "Файл не найден."

    group_name = group.upper()
    scopes = ["https://www.googleapis.com/auth/drive"]
    SERVICE_ACCOUNT_FILE = (
        "/home/cusdeb/Projects/lizard_bot/lizardbot/lizardbot-423509-18b41a862983.json"
    )
    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=scopes
    )
    drive_service = build("drive", "v3", credentials=credentials)

    file_content = download_file(chosen_file["id"], drive_service)
    results = process_excel(file_content, group_name)

    if not results:
        return "Неправильно введен номер группы или занятий нет."

    message = []
    for sheet_title, room_number, teacher_name, class_hour in results:
        if isinstance(room_number, float):
            room_number = int(room_number)
        if class_hour:
            message.append(
                f"\n{sheet_title},🔑 Кабинет: {room_number},💼 Преподаватель: {teacher_name}\n"
            )
        else:
            message.append(
                f"\n{sheet_title},🔑 Кабинет: {room_number},💼 Преподаватель: {teacher_name}\n"
            )
    message2 = f"{group_name.upper()}\n" + "".join(message).replace(",", "\n")
    message3 = form_schedule(message2)
    return message3


def search_schedule_by_teacher(name, teacher_name):
    """
    Обрабатывает запрос на получение расписания для указанного преподавателя.

    Аргументы:
        name (str): Название файла с расписанием (без расширения).
        teacher_name (str): Имя преподавателя для поиска.

    Возвращает:
        str: Отформатированное расписание или сообщение об ошибке.
    """
    files = get_filenames()

    chosen_file_name = name
    chosen_file = None
    for file in files:
        if file["name"] == chosen_file_name + ".xlsx":
            chosen_file = file
            break

    if chosen_file is None:
        return "Файл не найден."

    scopes = ["https://www.googleapis.com/auth/drive"]
    SERVICE_ACCOUNT_FILE = "/lizardbot/lizardbot-423509-18b41a862983.json"
    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=scopes
    )
    drive_service = build("drive", "v3", credentials=credentials)

    file_content = download_file(chosen_file["id"], drive_service)
    results = process_excel2(file_content, teacher_name)

    if not results:
        return "Неправильно введены данные или занятий нет."

    message = []
    for sheet_title, room_number, group_name, full_teacher_name, class_hour in results:
        if isinstance(room_number, float):
            room_number = int(room_number)
        if class_hour:
            message.append(
                f"\n{sheet_title},🔑 Кабинет: {room_number},💼 Группа: {group_name}\n"
            )
        else:
            message.append(
                f"\n{sheet_title},🔑 Кабинет: {room_number},💼 Группа: {group_name}\n"
            )
    message2 = f"{teacher_name.capitalize()}\n" + "".join(message).replace(",", "\n")
    message3 = form_schedule(message2)
    return message3
